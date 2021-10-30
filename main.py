import json
import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from random import randint, choice
from time import sleep

def get_tuple_to_lowercase (input_tuple : tuple):   
    result_tuple = tuple(item.lower() if type(item) is str else item for item in input_tuple) 
    return result_tuple

def tuple_match (A, B):
    if len(A) <= len(B):
        return A == tuple(b for b in B if b in A)
    else:
        return B == tuple(a for a in A if a in B)

def get_row_indexes(sheet, keys_tuple: tuple, start_row=1, start_col=1):
    indexes = {}     
    current_row_index = 1                           
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col, max_col=sheet.max_column, values_only=True):
        row = get_tuple_to_lowercase(row)
        if tuple_match(keys_tuple, row):
            for key in keys_tuple:
                indexes[key] = row.index(key)
            return current_row_index, indexes
        current_row_index += 1

# не делает то, как называется, потому что в какой-то момент все шло по ипзде из-за двойного дампа, который мне не хотелось по-нормальному исправлять
def create_users_json(sheet, header):
    users = {}
    user = {}
    user_id = 0
    header_row, user_header_indexes = get_row_indexes(sheet, header) 
    start_row = header_row + 1           # первая строчка после шапки    
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        row = get_tuple_to_lowercase(row)
        for column_name in header:
            user[column_name] = row[user_header_indexes[column_name]]
        user['completed'] = False
        users[user_id] = user.copy()
        user_id += 1 
    return users
            

def main():
    CURRENT_DIR = os.path.realpath(os.getcwd())
    # в files могут быть разные варианты ответов, они хранятся в json_files_list
    JSON_FILE_PATH = 'files/'
    json_files_path = os.path.realpath(os.path.join(os.getcwd(), JSON_FILE_PATH))
    json_files_list = []
    for File in os.listdir(json_files_path):
        if File.endswith('.json'):
            json_files_list.append(os.path.realpath(os.path.join(json_files_path, File)))

    # слишком тупой, чтобы починить это
    # в папке с main будет файл с информацией о пользователях
    # если его нет, он создается из xlsx файла, который называется всегда одинаково
    json_exists = False
    for File in os.listdir(CURRENT_DIR):
        if File.endswith('.json'):
            json_exists = True
            break
    if not json_exists:
        header = ('логин','пароль','пол', 'возраст')
        FILENAME = '392515-0019-7e.xlsx' # добавить аргумент скрипта | единственный файл 
        workbook = load_workbook(filename=FILENAME, read_only=True, data_only=True)
        sheet = workbook.active
        users_json = create_users_json(sheet, header)
        with open('users.json' , 'w', encoding='utf-8') as outfile:
            json.dump(users_json, outfile)
    for File in os.listdir(CURRENT_DIR):
        if File.endswith('.json'):
            json_exists = True
            break
    with open(File, 'r', encoding='utf-8') as json_file:
        users_json_dict = json.load(json_file)
        
    # web scraping 
    options = Options()
    options.headless = False
    options.add_argument("--window-size=800,600")
    DRIVER_NAME = 'chromedriver.exe'
    PAGE_NAME = 'http://39.soctest.ru'
    DRIVER_PATH = os.path.realpath(os.path.join(os.getcwd(), DRIVER_NAME))
    driver = webdriver.Chrome(options = options, executable_path=DRIVER_PATH) 
    driver.get(PAGE_NAME)

    for user_id in users_json_dict:
        if users_json_dict[user_id]['completed']:
            continue
        # для каждого пользователя выбирается случайный файл с ответами
        with open(choice(json_files_list) , 'r', encoding='utf-8') as json_file:
                questions_dict = json.load(json_file)
        while True:
            # начальный экран
            login_inputs = driver.find_elements_by_xpath('//*[@id="test_user_login"]')
            if login_inputs:
                login_inputs[0].send_keys(users_json_dict[user_id]['логин'])
            password_inputs = driver.find_elements_by_xpath('//*[@id="test_user_password"]')
            if login_inputs:
                password_inputs[0].send_keys(users_json_dict[user_id]['пароль'])
            enter_buttons = driver.find_elements_by_xpath('//button[contains(text(), "Войти")]')
            if enter_buttons:
                enter_buttons[0].click()

            #мы на регистрации, если есть кнопка сохранить
            save_buttons = driver.find_elements_by_xpath('//input[@value="Сохранить"]')
            if save_buttons:
                sex_input = driver.find_elements_by_xpath('//*[@id="test_user_sex"]')
                if sex_input:
                    sex_option_male = driver.find_element_by_xpath('//*[@id="test_user_sex"]/option[2]')
                    if users_json_dict[user_id]['пол'] == 'м':
                        sex_option_male.click()
                age_input = driver.find_elements_by_xpath('//*[@id="test_user_age"]')
                if age_input:
                    age_input[0].send_keys(users_json_dict[user_id]['возраст'])
                save_buttons[0].click()

            # какая-то мышиная возня с кнопками 
            # кнопки могут быть частью формы, тогда называются button, могут быть просто дивом класса btn (fake_button)
            buttons = driver.find_elements_by_xpath('//button')
            if len(buttons) == 1:
                if buttons[0].text == 'Выйти из системы':
                    users_json_dict[user_id]['completed'] = True # пользователь прошел тест
                    buttons[0].click
                    #записать изменения в файл
                    json_file = open('users.json', 'w')
                    json.dump(users_json_dict, json_file)
                    json_file.close()
                    break
                else:
                    buttons[0].click()
            elif len(buttons) > 1:
                # проверка ответов
                question_field = driver.find_element_by_xpath('//*[@id="ko-view"]/p/span')
                yes_button = driver.find_element_by_xpath('//button[contains(text(), "ДА")]')
                yyes_button = driver.find_element_by_xpath('//button[contains(text(), "Скорее ДА, чем НЕТ")]')
                no_button = driver.find_element_by_xpath('//button[contains(text(), "НЕТ")]')
                nno_button = driver.find_element_by_xpath('//button[contains(text(), "Скорее НЕТ, чем ДА")]')
                continue_button = driver.find_element_by_xpath('//button[contains(text(), "Ответить")]')
                try:
                    for question in questions_dict['questions']:
                        if question['content'] == question_field.text:
                            if question['answer'] == 'yes':
                                yes_button.click()
                            if question['answer'] == 'yyes':
                                yyes_button.click()
                            if question['answer'] == 'no':
                                no_button.click()
                            if question['answer'] == 'nno':
                                nno_button.click()
                            questions_dict['questions'].remove(question)
                            continue_button.click()
                            sleep(randint(3,5))
                except:
                    button = driver.find_element_by_xpath('//a[contains(@class, "btn")]')
                    button.click()
            # копипаста предыдущих 40 строк
            fake_buttons = driver.find_elements_by_xpath('//a[contains(@class, "btn")]')
            if len(fake_buttons) == 1:
                if fake_buttons[0].text == 'Выйти из системы':
                    users_json_dict[user_id]['completed'] = True # пользователь прошел тест
                    fake_buttons[0].click()
                    #записать изменения в файл
                    json_file = open('users.json', 'w')
                    json.dump(users_json_dict, json_file)
                    json_file.close()
                    break
                else:
                    fake_buttons[0].click()
            elif len(fake_buttons) > 1:
                # проверка ответов
                question_field = driver.find_element_by_xpath('//*[@id="ko-view"]/p/span')
                yes_button = driver.find_element_by_xpath('//button[contains(text(), "ДА")]')
                yyes_button = driver.find_element_by_xpath('//button[contains(text(), "Скорее ДА, чем НЕТ")]')
                no_button = driver.find_element_by_xpath('//button[contains(text(), "НЕТ")]')
                nno_button = driver.find_element_by_xpath('//button[contains(text(), "Скорее НЕТ, чем ДА")]')
                continue_button = driver.find_element_by_xpath('//button[contains(text(), "Ответить")]')
                for question in questions_dict['questions']:
                    if question['content'] == question_field.text:
                        if question['answer'] == 'yes':
                            yes_button.click()
                        if question['answer'] == 'yyes':
                            yyes_button.click()
                        if question['answer'] == 'no':
                            no_button.click()
                        if question['answer'] == 'nno':
                            nno_button.click()
                        questions_dict['questions'].remove(question)
                        continue_button.click()
                        sleep(randint(3,7))
    driver.quit()   

    
if __name__ == '__main__':
    main()
